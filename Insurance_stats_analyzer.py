import os
import pandas as pd
import xlrd
import re
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys

class Logger:
    def __init__(self, filename):
        self.terminal = sys.stdout
        self.log = open(filename, "w")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass

today = datetime.today().strftime('%Y-%m-%d')
sys.stdout = Logger(f"Formulas for the file Comparison of IC {today}.txt")

# your entire code

path = r"CASCO IC new 2024\9 months"

# Path to the folder containing subfolders
files_and_folders = os.listdir(path)

# Sheet names with reports in the corresponding folders
def direct_sheet_name(sheet_names):
    filter_sheet = [sheet for sheet in sheet_names if "2.3" in sheet or "2_3" in sheet]
    if len(filter_sheet) == 1:
        return filter_sheet[0]
    if len(filter_sheet) > 1:
        for sheet in filter_sheet:
            if 'Без' in sheet:
                return sheet
    return sheet_names[0]

# Names of the required columns selection
def find_matching_sheet(df, patterns_list):
    for patterns in patterns_list:
        if all(any(pattern in col for col in df.columns) for pattern in patterns):
            df_filtered = df[[col for col in df.columns if any(pattern in col for pattern in patterns)]]
            return df_filtered

patterns_list = [
    ["Наименование показателя", "1 - добровольное", "2.1 - ", "3 - обязательное", "7 - страхование", "16 - "], 
    ["Перечень учетных", "1 - добровольное", "2.1 - ", "3 - обязательное", "7 - страхование", "16 - "],
    ["Наименование показателя", "групп 1 ", "групп 2 ", "групп 3 ", "групп 7 ", "групп 16 "],
    ["Наименование показателя", " группа № 1 ", "группа № 2.1 ", "группа № 3 ", "группа № 7 ", "группа № 16 "],
    ["Наименование показателя", "группа 1.", "группа 2.1.", "группа 3.", "группа 7.", "группа 16."],
    ["Перечень учетных групп", "1 - ", "2.1 - "]]

dict_str = {
    1.0: 'Заработанные страховые премии – нетто-перестрахование, в том числе:',
    1.1: 'Страховые премии по операциям страхования, сострахования и перестрахования – нетто-перестрахование по страхованию иному, чем страхование жизни',
    1.2: 'Изменение резерва незаработанной премии – нетто-перестрахование по страхованию иному, чем страхование жизни',
    2.0: 'Состоявшиеся убытки – нетто-перестрахование, в том числе:',
    2.1: 'Выплаты по операциям страхования, сострахования и перестрахования – нетто-перестрахование по страхованию иному, чем страхование жизни',
    2.2: 'Расходы по урегулированию убытков – нетто-перестрахование по страхованию иному, чем страхование жизни',
    2.3: 'Изменение резервов убытков – нетто-перестрахование по страхованию иному, чем страхование жизни',
    2.4: 'Доходы от регрессов, суброгаций и прочих возмещений – нетто-перестрахование по страхованию иному, чем страхование жизни',
    2.5: 'Изменение оценки будущих поступлений по регрессам, суброгациям и прочим возмещениям – нетто-перестрахование по страхованию иному, чем страхование жизни',
    3.0: 'Расходы по ведению страховых операций – нетто-перестрахование, в том числе:',
    3.1: 'Аквизиционные расходы – нетто-перестрахование (по договорам страхования иного, чем страхование жизни)',
    3.2: 'Изменение отложенных аквизиционных расходов и доходов (по договорам страхования иного, чем страхование жизни)',
    4.0: 'Отчисления от страховых премий',
    5.0: 'Прочие доходы по страхованию иному, чем страхование жизни',
    6.0: 'Прочие расходы по страхованию иному, чем страхование жизни',
    7.0: 'Результат от операций по страхованию иному, чем страхование жизни'
}

# Renaming of indicator names
dict_str_2 = {
    'Заработанные страховые премии – нетто-перестрахование': dict_str[1.0],
    'Страховые премии по операциям страхования, сострахования и перестрахования – нетто-перестрахование': dict_str[1.1],
    'Изменение резерва незаработанной премии – нетто-перестрахование': dict_str[1.2],
    'Состоявшиеся убытки – нетто-перестрахование': dict_str[2.0],
    'Выплаты по операциям страхования, сострахования и перестрахования – нетто-перестрахование': dict_str[2.1],
    'Расходы по урегулированию убытков – нетто-перестрахование': dict_str[2.2],
    'Изменение резервов убытков – нетто-перестрахование': dict_str[2.3],
    'Доходы от регрессов, суброгаций и прочих возмещений – нетто-перестрахование': dict_str[2.4],
    'Изменение оценки будущих поступлений по регрессам, суброгациям и прочим возмещениям – нетто-перестрахование': dict_str[2.5],
    'Расходы по ведению страховых операций – нетто-перестрахование': dict_str[3.0],
    'Аквизиционные расходы – нетто-перестрахование':  dict_str[3.1],
    'Изменение отложенных аквизиционных расходов и доходов': dict_str[3.2],
    'Отчисления от страховых премий': dict_str[4.0],
    'Прочие доходы по страхованию иному, чем страхование жизни – нетто-перестрахование': dict_str[5.0],
    'Прочие расходы по страхованию иному, чем страхование жизни – нетто-перестрахование': dict_str[6.0],
    'Результат от операций по страхованию иному, чем страхование жизни – нетто-перестрахование': dict_str[7.0]
}

# Dictionary with corresponding stat values and their tables
dfs = {}
i = 0

# Cleaning of sheet names
def clean_column_name(name):
    if isinstance(name, str):
        name = re.sub(r"Unnamed:\s*\d+_level_\d+", '', name)
        name = re.sub(r'\n+','', name)
        name = re.sub(r'\s+\s+',' ', name)
        return name.strip()
    return name
    
# Standardization of required column names
def raname_columns(col_name):
    if ((('Перечень учетных групп 16 'in col_name) or ('учетная группа 16. ' in col_name) or ('Учетная группа № 16 ' in col_name)) and 'отчетный период' in col_name) or ('16 - страхование лиц, выезжающих за пределы постоянного места жительства отчетный период' in col_name):
        return 'ВЗР 16 отчетный период'
    elif ((('Перечень учетных групп 16 ' in col_name) or ('учетная группа 16. ' in col_name) or ('Учетная группа № 16 ' in col_name)) and 'аналогичный период предыдущего года' in col_name) or ('16 - страхование лиц, выезжающих за пределы постоянного места жительства аналогичный период предыдущего года' in col_name):
        return 'ВЗР 16 период предыдущего года'
    elif ((('Перечень учетных групп 7 ' in col_name) or ('учетная группа 7. ' in col_name) or ('Учетная группа № 7' in col_name)) and 'отчетный период' in col_name) or ('7 - страхование средств наземного транспорта отчетный период' in col_name):
        return 'КАСКО 7 отчетный период'
    elif ((('Перечень учетных групп 7 ' in col_name) or ('учетная группа 7. ' in col_name) or ('Учетная группа № 7' in col_name)) and 'аналогичный период предыдущего года' in col_name) or ('7 - страхование средств наземного транспорта аналогичный период предыдущего года' in col_name):
        return 'КАСКО 7 период предыдущего года'
    elif ((('Перечень учетных групп 3 ' in col_name) or ('учетная группа 3. ' in col_name) or ('Учетная группа № 3 ' in col_name)) and 'отчетный период' in col_name) or ('3 - обязательное страхование гражданской ответственности владельцев транспортных средств отчетный период' in col_name):
        return 'ОСАГО 3 отчетный период'
    elif ((('Перечень учетных групп 3 ' in col_name) or ('учетная группа 3. ' in col_name) or ('Учетная группа № 3 ' in col_name)) and 'аналогичный период предыдущего года' in col_name) or ('3 - обязательное страхование гражданской ответственности владельцев транспортных средств аналогичный период предыдущего года' in col_name):
        return 'ОСАГО 3 период предыдущего года'
    elif ((('Перечень учетных групп 2' in col_name) or ('учетная группа 2.' in col_name) or ('Учетная группа № 2.1 ' in col_name)) and 'отчетный период' in col_name) or ('2.1 - страхование от несчастных случаев и болезней отчетный период' in col_name):
        return 'HC 2 отчетный период'
    elif ((('Перечень учетных групп 2' in col_name) or ('учетная группа 2.' in col_name) or ('Учетная группа № 2.1 ' in col_name)) and 'аналогичный период предыдущего года' in col_name) or ('2.1 - страхование от несчастных случаев и болезней аналогичный период предыдущего года' in col_name):
        return 'HC 2 период предыдущего года'
    elif ((('Перечень учетных групп 1 ' in col_name) or ('учетная группа 1. ' in col_name) or ('Учетная группа № 1 ' in col_name)) and 'отчетный период'in col_name) or ('1 - добровольное медицинское страхование отчетный период'  in col_name):
        return 'ДМС 1 отчетный период'
    elif ((('Перечень учетных групп 1 ' in col_name) or ('учетная группа 1. ' in col_name) or ('Учетная группа № 1 ' in col_name)) and 'аналогичный период предыдущего года' in col_name) or ('1 - добровольное медицинское страхование аналогичный период предыдущего года'  in col_name):
        return 'ДМС 1 период предыдущего года'
    else:
        return col_name

# Function for calculating stats
def apply_value(row, column_name, df):
    try:
        # Check if the value exists before accessing it
        if row['Наименование показателя'] == 'НП':
            value = df.loc[df['Наименование показателя'] == 
                           'Страховые премии по операциям страхования, сострахования и перестрахования – нетто-перестрахование по страхованию иному, чем страхование жизни', column_name]
            value = value.values[0] if not value.empty else 0
            print(f'НП: {value}')        
            return value
        
        elif row['Наименование показателя'] == 'ЗП':
            value = df.loc[df['Наименование показателя'] == 
                           'Заработанные страховые премии – нетто-перестрахование, в том числе:', column_name]
            value = value.values[0] if not value.empty else 0
            print(f'ЗП: {value}') 
            return value if not pd.isna(value) else 0
        
        elif row['Наименование показателя'] == 'КУ':
            payments = df.loc[df['Наименование показателя'] == 
                              'Выплаты по операциям страхования, сострахования и перестрахования – нетто-перестрахование по страхованию иному, чем страхование жизни', column_name]
            reserves_change = df.loc[df['Наименование показателя'] == 
                                     'Изменение резервов убытков – нетто-перестрахование по страхованию иному, чем страхование жизни', column_name]
            ZP = df.loc[df['Наименование показателя'] == 
                        'Заработанные страховые премии – нетто-перестрахование, в том числе:', column_name]
            payments = payments.values[0] if not payments.empty else 0
            reserves_change = reserves_change.values[0] if not reserves_change.empty else 0
            ZP = ZP.values[0] if not ZP.empty else 1  # Prevent division by zero
            value = -(payments + reserves_change) / ZP
            print(f'КУ: -({payments}+{reserves_change})/{ZP}') 
            return value if not pd.isna(value) else 0
        
        elif row['Наименование показателя'] == 'РУУ с регр и субр':
            expenses = df.loc[df['Наименование показателя'] == 
                             'Расходы по урегулированию убытков – нетто-перестрахование по страхованию иному, чем страхование жизни', column_name]
            income_from_recourses = df.loc[df['Наименование показателя'] == 
                                           'Доходы от регрессов, суброгаций и прочих возмещений – нетто-перестрахование по страхованию иному, чем страхование жизни', column_name]
            change_in_rating = df.loc[df['Наименование показателя'] == 
                                      'Изменение оценки будущих поступлений по регрессам, суброгациям и прочим возмещениям – нетто-перестрахование по страхованию иному, чем страхование жизни', column_name]
            ZP = df.loc[df['Наименование показателя'] == 
                        'Заработанные страховые премии – нетто-перестрахование, в том числе:', column_name]
            expenses = expenses.values[0] if not expenses.empty else 0
            income_from_recourses = income_from_recourses.values[0] if not income_from_recourses.empty else 0
            change_in_rating = change_in_rating.values[0] if not change_in_rating.empty else 0
            ZP = ZP.values[0] if not ZP.empty else 1  # Prevent division by zero
            value = -(expenses + income_from_recourses + change_in_rating) / ZP
            print(f'РУУ с регр и субр: -({expenses}+{income_from_recourses}+{change_in_rating})/{ZP}') 
            return value if not pd.isna(value) else 0
        
        elif row['Наименование показателя'] == 'Аквизиция':
            expenses = df.loc[df['Наименование показателя'] == 
                             'Расходы по ведению страховых операций – нетто-перестрахование, в том числе:', column_name]
            ZP = df.loc[df['Наименование показателя'] == 
                        'Заработанные страховые премии – нетто-перестрахование, в том числе:', column_name]
            expenses = expenses.values[0] if not expenses.empty else 0
            ZP = ZP.values[0] if not ZP.empty else 1  # Prevent division by zero
            value = -expenses / ZP
            print(f'Аквизиция: -({expenses})/{ZP}') 
            return value if not pd.isna(value) else 0
        
        elif row['Наименование показателя'] == 'ККУ':
            payments = df.loc[df['Наименование показателя'] == 
                             'Результат от операций по страхованию иному, чем страхование жизни', column_name]
            ZP = df.loc[df['Наименование показателя'] == 
                        'Заработанные страховые премии – нетто-перестрахование, в том числе:', column_name]
            payments = payments.values[0] if not payments.empty else 0
            ZP = ZP.values[0] if not ZP.empty else 1  # Prevent division by zero
            value = 1 - payments / ZP
            print(f'ККУ: {1 - payments}/{ZP}') 
            return value if not pd.isna(value) else 0
        
    except Exception as e:
        print(f"Error in apply_value for {row['Наименование показателя']}: {e}")
        return 0

for folder in files_and_folders:
    folder_path = os.path.join(path, folder)  # Opening the folder
    if os.path.isdir(folder_path):
        files = os.listdir(folder_path)
        for file in files:
            if "0420158" in file:  # Opening the file with these numbers
                if file.endswith(".xls") or file.endswith(".xlsx"):  # Depending on the extension
                    file_path = os.path.join(folder_path, file)
                    if file.endswith(".xls"):
                        sheet_names = pd.ExcelFile(file_path, engine='xlrd').sheet_names
                        sheet_name = direct_sheet_name(sheet_names)  # Page selected
                        if folder == "Согласие":
                            df = pd.read_excel(file_path, sheet_name=sheet_name, header=[1, 2], engine='xlrd') 
                        else:
                            df = pd.read_excel(file_path, sheet_name=sheet_name, header=[2, 3, 4, 5], engine='xlrd') 
                    elif file.endswith(".xlsx"):
                        sheet_names = pd.ExcelFile(file_path).sheet_names
                        sheet_name = direct_sheet_name(sheet_names)  # Выбрали страницу 
                        if folder in ["Тинькофф", "Т-страхование (бывш. Тинькоф)", "Т-страхование", "СОГАЗ", "Ренессанс", "Согаз"]:
                            df = pd.read_excel(file_path, sheet_name=sheet_name, header=[1, 2, 3])
                        else:
                            df = pd.read_excel(file_path, sheet_name=sheet_name, header=[2, 3, 4, 5])

                    # Cleaning column names
                    df.columns = [" ".join(map(str, col)).strip() for col in df.columns.values]
                    df.columns = [clean_column_name(col) for col in df.columns] 

                    # Removing unnecessary and empty rows
                    if (df.isnull().all(axis=1).idxmax() > 15):
                        empty_row_idx = df.isnull().all(axis=1).idxmax()
                        df = df.loc[:empty_row_idx-1]
                    df = df.dropna(axis=1, how='all')

                    # Selecting only the necessary columns
                    df_filtered = find_matching_sheet(df, patterns_list)
                    # Deleting empty columns
                    df_filtered = df_filtered.loc[:, ~df_filtered.columns.duplicated()]
                    
                    # Renaming the first column
                    if 'Перечень учетных групп' in df_filtered.columns:
                        df_filtered.rename(columns={'Перечень учетных групп': 'Наименование показателя'}, inplace=True)
                        
                    # Cleaning values
                    for col in df_filtered.columns.difference(['Наименование показателя']):
                        # Replacing round brackets with a hyphen in all column values
                        df_filtered[col] = df_filtered[col].replace(['-', 'X', 'x', '*'], 0.0)
                        df_filtered[col] = df_filtered[col].apply(lambda x: str(x).replace('(', '-').replace(')', '').replace(',', '.').replace(' ', '')).astype(float)
                        df_filtered[col] = df_filtered[col].replace(['0.0', '0', 'nan'], 0.0).astype(float)
                        if folder in ["Ресо", "Согласие"]:
                            df_filtered[col] = df_filtered[col].mul(1000)
                            
                    # Removing duplicate columns
                    if df_filtered.columns.duplicated().sum() > 0:
                        df_filtered = df_filtered.loc[:, ~df_filtered.columns.duplicated()]

                    # Unify the first column
                    df_filtered['Наименование показателя'] = df_filtered['Наименование показателя'].replace(dict_str)
                    df_filtered['Наименование показателя'] = df_filtered['Наименование показателя'].str.strip()
                    df_filtered['Наименование показателя'] = df_filtered['Наименование показателя'].replace(r'\s+', ' ', regex=True)
                    df_filtered['Наименование показателя'] = df_filtered['Наименование показателя'].str.capitalize()
                    df_filtered['Наименование показателя'] = df_filtered['Наименование показателя'].replace(dict_str_2)
                    df_filtered = df_filtered.loc[:, (df_filtered != 0).any(axis=0)]

                    # Renaming columns
                    df_filtered.columns = [raname_columns(col) for col in df_filtered.columns]
                    # Arrange in ascending order
                    sorted_columns = sorted(df_filtered.columns, key=lambda x: (re.findall(r'\d+', x), x))
                    df_filtered = df_filtered[sorted_columns]

                    # Replace empty values with 0 for further calculations
                    df_filtered = df_filtered.fillna(0)

                    # Assign a name to the table
                    if folder == "Сбербанк Страхование":
                        globals()[f"df_Сбербанк_Страхование"] = df_filtered
                        dfs[folder] = df_filtered
                    elif folder == "Согласие-Вита":
                        df_filtered = df_filtered.drop(df_filtered.index[0])
                        df_filtered = df_filtered.reset_index(drop=True)
                        globals()[f"df_Согласие_Вита"] = df_filtered
                        dfs[folder] = df_filtered
                    elif folder == "Согласие":
                        df_filtered = df_filtered.drop(df_filtered.index[0])
                        df_filtered = df_filtered.reset_index(drop=True)
                        globals()[f"df_Согласие"] = df_filtered
                        dfs[folder] = df_filtered
                    elif folder == "Т-страхование (бывш. Тинькоф)":
                        globals()[f"df_Т_страхование"] = df_filtered
                        dfs[folder] = df_filtered
                    else:
                        globals()[f"df_{folder}"] = df_filtered
                        dfs[folder] = df_filtered
                    i += 1

# Creating dataframes by companies
stat = {}
i = 0
# Calculate stats
for name, df in dfs.items():
    new_rows = pd.DataFrame({'Наименование показателя': ['НП', 'ЗП', 'КУ', 'РУУ с регр и субр', 'Аквизиция', 'ККУ']})
    for col in df.columns[1:]:
        new_rows[col] = new_rows.apply(lambda row: apply_value(row, col, df), axis=1)
    globals()[name] = new_rows
    stat[name] = new_rows
    i += 1

# Creating dataframes by business lines
categories = ["ДМС", "HC", "ОСАГО", "КАСКО", "ВЗР"]
page = {}
i = 0

for category in categories:
    new_rows = pd.DataFrame({'Наименование показателя': ['НП', 'ЗП', 'КУ', 'РУУ с регр и субр', 'Аквизиция', 'ККУ']})
    for df_name, df in stat.items():
        df_to_list = df.loc[:, [col for col in df.columns if category in col]]
        df_to_list.columns = [f"{col} {df_name}" for col in df_to_list.columns]
        if category not in globals():
            globals()[category] = df_to_list
        else:
            globals()[category] = pd.concat([globals()[category], df_to_list], axis=1)
        i += 1
    result = pd.concat([new_rows, globals()[category]], axis=1)
    page[f"{category}"] = result

# Create a new Excel file
file_name = f'Comparison SK {today}.xlsx'
wb = Workbook()

# Formats for borders and fonts
bold_font = Font(bold=True)
center_align = Alignment(horizontal='center', vertical='center') 
border_style = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Function to apply cell formatting
def format_cells(ws, range_start, range_end):
    for row in ws.iter_rows(min_row=range_start[0], max_row=range_end[0],
                            min_col=range_start[1], max_col=range_end[1]):
        for cell in row:
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border_style

# Function to auto adjust column width
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Determine the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Add a bit to the width
        ws.column_dimensions[col_letter].width = adjusted_width

# Iterate through each category and create separate sheets
for category, df in page.items():
    ws = wb.create_sheet(title=category)

    # Add headers 'Company' and 'Period'
    ws.cell(row=1, column=1, value='Company')
    ws.cell(row=2, column=1, value='Period')

    # Column index starts from the 2nd (the first column is 'Company' and 'Period')
    col_idx = 2

    # Add headers for each company and periods
    for col in df.columns[1:]:
        company_name = col.split(' ')[-1]  # Get the company name
        period_type = 'reporting period' if 'reporting period' in col else "similar period of the previous year"

        # Check if cells should be merged for the same companies
        if col_idx > 2 and ws.cell(row=1, column=col_idx - 1).value == company_name:
            # Merge cells with the company name if they match
            ws.merge_cells(start_row=1, start_column=col_idx-1, end_row=1, end_column=col_idx)
        else:
            ws.cell(row=1, column=col_idx, value=company_name)  # Write the company name

        # Fill in labels for "Reporting period" and "Similar period"
        ws.cell(row=2, column=col_idx, value=period_type)
        col_idx += 1

    # Fill data starting from the 3rd row
    for i, row in df.iterrows():
        ws.append([row['Indicator Name']] + row.tolist()[1:])

    # Apply formatting to the first row and column
    format_cells(ws, (1, 1), (2, len(df.columns)))  # For the first two rows
    format_cells(ws, (3, 1), (ws.max_row, len(df.columns)))  # For all rows with data

    # Apply number format
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        first_col = row[0].value
        if first_col in ['НП', 'ЗП']:  # Apply to specific columns
            for cell in row[1:]:
                cell.number_format = '#,##0'  # Numeric format
                cell.font = Font(bold=False)
        else:
            for cell in row[1:]:
                cell.number_format = '#,##0.0%'  # Percentage format
                cell.font = Font(bold=False)
        cell.border = border_style

    # Apply bold formatting to the first row and first column
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = bold_font
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).font = bold_font

    # Auto-adjust column width
    auto_adjust_column_width(ws)

# Remove default empty first sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Save the file
wb.save(file_name)
